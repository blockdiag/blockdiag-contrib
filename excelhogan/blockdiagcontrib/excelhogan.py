# -*- coding: utf-8 -*-
#  Copyright 2012 Takeshi KOMIYA
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

import Image
import ImageDraw
import openpyxl
from openpyxl.style import Color, Fill
from openpyxl.cell import get_column_letter
from blockdiag.imagedraw import base, textfolder
from blockdiag.imagedraw.png import ttfont_for
from blockdiag.utils import Size, XY
from blockdiag.utils.fontmap import FontMap

import patch  # enable zoomScale feature


def dots_of_line(pt1, pt2):
    if pt1.x == pt2.x:  # vertical
        if pt1.y > pt2.y:
            pt2, pt1 = pt1, pt2

        for y in xrange(pt1.y, pt2.y + 1):
            yield XY(pt1.x, y)

    elif pt1.y == pt2.y:  # horizontal
        if pt1.x > pt2.x:
            pt2, pt1 = pt1, pt2

        for x in xrange(pt1.x, pt2.x + 1):
            yield XY(x, pt1.y)
    else:  # diagonal
        if pt1.x > pt2.x:
            pt2, pt1 = pt1, pt2

        # DDA (Digital Differential Analyzer) Algorithm
        m = float(pt2.y - pt1.y) / float(pt2.x - pt1.x)
        x = pt1.x
        y = pt1.y

        while x <= pt2.x + 1:
            yield XY(int(x), int(round(y)))
            x += 1
            y += m


def text_bitmap(string, size, font):
    ttfont = ttfont_for(font)
    draw = ExcelHoganImageDraw(None)
    draw.set_canvas_size(size)
    draw.drawer.text((0, 0), string, font=ttfont, fill='black')

    bitmap = [[] for _ in range(size.width)]
    for i, pixel in enumerate(draw.image.getdata()):
        if pixel[-1]:
            data = 1
        else:
            data = 0

        bitmap[i % size.width].append(data)

    return bitmap


class ExcelHoganImageDraw(base.ImageDraw):
    def __init__(self, filename, **kwargs):
        self.filename = filename
        self.image = None
        self.drawer = None

    def set_canvas_size(self, size):
        self.image = Image.new('RGBA', size)
        self.drawer = ImageDraw.Draw(self.image)

        self.book = openpyxl.Workbook()
        self.sheet = self.book.worksheets[0]

        # setup default zoom of sheet
        sheet_view = openpyxl.worksheet.SheetView()
        sheet_view.zoomScale = "10"
        sheet_view.zoomScaleNormal = "10"
        self.sheet.sheet_view = sheet_view

        # setup sheet as HOGAN
        coldim = openpyxl.worksheet.ColumnDimension()
        coldim.width = 4

        for i in range(1, 26 * 26):
            colname = get_column_letter(i)
            self.sheet.column_dimensions[colname] = coldim

    def point(self, pt, fill):
        color = "00%02X%02X%02X" % fill

        cell = self.sheet.cell(column=pt[0], row=pt[1])
        cell.style.fill.fill_type = Fill.FILL_SOLID
        cell.style.fill.start_color = Color(color)
        cell.style.fill.end_color = Color(color)

    def rectangle(self, box, **kwargs):
        fill = kwargs.get('fill')
        if fill:
            for x in range(box.x1, box.x2 + 1):
                for y in range(box.y1, box.y2 + 1):
                    self.point((x, y), fill=fill)

        outline = kwargs.get('outline')
        if outline:
            self.line((XY(box.x1, box.y1), XY(box.x2, box.y1)), fill=outline)
            self.line((XY(box.x1, box.y2), XY(box.x2, box.y2)), fill=outline)
            self.line((XY(box.x1, box.y1), XY(box.x1, box.y2)), fill=outline)
            self.line((XY(box.x2, box.y1), XY(box.x2, box.y2)), fill=outline)

    def line(self, xy, **kwargs):
        fill = kwargs.get('fill')

        for pt in dots_of_line(*xy):
            self.point(pt, fill=fill)

    def textsize(self, string, font, maxwidth=None, **kwargs):
        if maxwidth is None:
            maxwidth = 65535

        box = (0, 0, maxwidth, 65535)
        textbox = textfolder.get(self, box, string, font, **kwargs)
        return textbox.outlinebox.size

    def textlinesize(self, string, font, **kwargs):
        ttfont = ttfont_for(font)
        if ttfont is None:
            size = self.drawer.textsize(string, font=None)

            font_ratio = font.size * 1.0 / FontMap.BASE_FONTSIZE
            size = Size(int(size[0] * font_ratio),
                        int(size[1] * font_ratio))
        else:
            size = self.drawer.textsize(string, font=ttfont)
            size = Size(*size)

        return size

    def text(self, xy, string, font, **kwargs):
        size = self.textlinesize(string, font)

        fill = kwargs.get('fill')
        for dx, line in enumerate(text_bitmap(string, size, font)):
            for dy, pixel in enumerate(line):
                if pixel:
                     self.point(xy.shift(dx, dy), fill=fill)

    def textarea(self, box, string, font, **kwargs):
        lines = textfolder.get(self, box, string, font, **kwargs)
        for string, xy in lines.lines:
            self.text(xy, string, font, **kwargs)

    def polygon(self, points, **kwargs):
        fill = kwargs.get('fill')
        if fill:
            pass

        outline = kwargs.get('outline')
        if outline:
            for i in xrange(len(points)):
                self.line((points[i], points[i - 1]), fill=outline)

    def save(self, filename, size, format):
        self.book.save(self.filename)


def setup(self):
    from blockdiag.imagedraw import install_imagedrawer
    install_imagedrawer('hogan.xlsx', ExcelHoganImageDraw)
